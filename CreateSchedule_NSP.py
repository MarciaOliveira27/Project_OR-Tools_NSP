import openpyxl
from openpyxl.styles import PatternFill

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('Schedule')

schedule_page = book['Schedule']

solution_1 = [
    ['', 'Day 0', 'Day 1', 'Day 2', 'Day 3', 'Day 4', 'Day 5', 'Day 6'],
    ['Nurse 0', 'shift 1', 'does not work', 'does not work', 'does not work', 'shift 2', 'shift 1', 'shift 0'],
    ['Nurse 1', 'does not work', 'shift 0', 'does not work', 'does not work', 'shift 1', 'shift 1', 'shift 1'],
    ['Nurse 2', 'does not work', 'does not work', 'does not work', 'shift 2', 'shift 1', 'shift 2', 'shift 2'],
    ['Nurse 3', 'does not work', 'does not work', 'shift 2', 'does not work', 'shift 0', 'shift 0', 'shift 0'],
    ['Nurse 4', 'does not work', 'does not work', 'shift 1', 'does not work', 'shift 0', 'shift 0', 'shift 1'],
    ['Nurse 5', 'shift 2', 'shift 1', 'shift 2', 'shift 2', 'does not work', 'does not work', 'does not work'],
    ['Nurse 6', 'shift 0', 'shift 0', 'shift 0', 'shift 1', 'does not work', 'does not work', 'shift 2'],
    ['Nurse 7', 'shift 2', 'shift 1', 'shift 0', 'shift 1', 'does not work', 'does not work', 'does not work'],
    ['Nurse 8', 'shift 1', 'shift 2', 'shift 1', 'shift 0', 'does not work', 'does not work', 'does not work'],
    ['Nurse 9', 'shift 0', 'shift 2', 'does not work', 'shift 0', 'shift 2', 'shift 2', 'does not work'],
]

for row in solution_1:
    schedule_page.append(row)

for row in schedule_page.iter_rows(min_row=2): 
    for cell in row:
        if cell.value == 'does not work':
            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        elif cell.value == 'shift 0':
            cell.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')   
        elif cell.value == 'shift 1':
            cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        elif cell.value == 'shift 2':
            cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')


book.save('ScheduleExcel_NSP.xlsx')
